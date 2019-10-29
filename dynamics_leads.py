import os
import openpyxl
import sqlite3
import csv
import datetime
import shutil
import pyperclip


class Global:
    def __init__(self):
        self.excel_import_path = ""
        self.upload_file_path = ""
        self.letter_merge_path = ""
        self.database = 'dynamics-leads.db'

        self.ia_ne_counties = {'IA_HARRISON': 'OMA', 'IA_MILLS': 'OMA', 'IA_POTTAWATTAMIE': 'OMA',
                               'NE_BUTLER': 'OMA', 'NE_CASS': 'OMA', 'NE_DODGE': 'OMA', 'NE_DOUGLAS': 'OMA',
                               'NE_LANCASTER': 'OMA', 'NE_SARPY': 'OMA', 'NE_SAUNDERS': 'OMA',
                               'NE_WASHINGTON': 'OMA'}

        self.tracking_codes = {'ML6': '2020 ML Guide', 'FSI20 ML2': '2020 ML Guide',
                               'FSI20 ML7': '2020 Adv Sol MN', 'FSI20 ML5': '2020 ML Guide',
                               'FSI20 ML3': '2020 Adv Sol CHI', 'FSI20 ML4': '2020 Adv Sol CHI'}

        self.transaction_types = {'ML6': 'C', 'FSI20 ML2': 'C',
                                  'FSI20 ML7': 'AD', 'FSI20 ML5': 'C',
                                  'FSI20 ML3': 'AD', 'FSI20 ML4': 'AD'}

        self.mn_counties = {'ANOKA': 'TC-TCM', 'CARVER': 'TC-TCM', 'DAKOTA': 'TC-TCM', 'HENNEPIN': 'TC-TCM',
                            'RAMSEY': 'TC-TCM', 'SCOTT': 'TC-TCM', 'WASHINGTON': 'TC-TCM',
                            'CHISAGO': 'TC-GTCM', 'ISANTI': 'TC-GTCM', 'STEARNS': 'TC-GTCM',
                            'KANDIYOHI': 'TC-GTCM', 'WRIGHT': 'TC-GTCM', 'SHERBURNE': 'TC-GTCM',
                            'BLUE EARTH': 'TC-SEMN', 'BROWN': 'TC-SEMN', 'DODGE': 'TC-SEMN',
                            'FARIBAULT': 'TC-SEMN', 'FILLMORE': 'TC-SEMN', 'FREEBORN': 'TC-SEMN',
                            'HOUSTON': 'TC-SEMN', 'MARTIN': 'TC-SEMN', 'MOWER': 'TC-SEMN',
                            'NICOLLET': 'TC-SEMN', 'OLMSTED': 'TC-SEMN', 'STEELE': 'TC-SEMN',
                            'WABASHA': 'TC-SEMN', 'WASECA': 'TC-SEMN', 'WATONWAN': 'TC-SEMN',
                            'WINONA': 'TC-SEMN'}

        self.merge_letter_header = ['Campaign', 'Individual_First_Name_1', 'Individual_Last_Name_1',
                                    'Individual_First_Name_2', 'Individual_Last_Name_2', 'Address_1',
                                    'Address_2', 'City', 'State', 'Zip', 'County', 'Unique_ID',
                                    'mid', 'art_code', 'kit'
                                    ]

        self.dynamics_header = ['collateral_kit_order', 'row_checksum', 'modified_on_do_not_mod', 'name',
                                'created_on', 'created_by', 'owner', 'prospect_or_broker',
                                'street_1', 'street_2', 'city', 'full_state', 'zipcode',
                                'county', 'state', 'collateral_kit', 'second_kit',
                                'application_url', 'comments', 'vendor_unique_id_lead',
                                'vendor_unique_id_contact', 'quantity', 'routing',
                                'modified_by', 'modified_on', 'ship_date']

        self.outside_area_header = ['filename', 'recno', 'collateral_kit_order', 'row_checksum',
                                    'modified_on_do_not_mod', 'name',
                                    'created_on', 'created_by', 'owner', 'prospect_or_broker',
                                    'street_1', 'street_2', 'city', 'full_state', 'zipcode',
                                    'county', 'state', 'collateral_kit', 'second_kit',
                                    'application_url', 'comments', 'vendor_unique_id_lead',
                                    'vendor_unique_id_contact', 'quantity', 'routing',
                                    'modified_by', 'modified_on', 'ship_date']

    def initialize_config(self):
        self.excel_import_path = os.path.join(os.path.curdir, 'downloaded')
        self.letter_merge_path = os.path.join(os.path.curdir, 'letter_merge')
        self.upload_file_path = os.path.join(self.excel_import_path, 'upload')


def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d


def import_leads(fle):
    file_path = os.path.join(g.excel_import_path, fle)
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.active

    conn = sqlite3.connect(database=g.database)
    cursor = conn.cursor()

    for n, row in enumerate(ws.iter_rows()):
        row_data = [cell.value for cell in row]
        # print(row_data)

        sql = ("INSERT INTO `records` VALUES ("
               "?,?,DATETIME('now', 'localtime'),"
               "?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);")

        if n != 0:
            cursor.execute(sql, (fle, n, None, None, row_data[0], row_data[1], row_data[2],
                                 row_data[3], row_data[4], row_data[5], row_data[6],
                                 row_data[7], row_data[8], row_data[9], row_data[10],
                                 row_data[11], row_data[12], row_data[13], row_data[14],
                                 row_data[15], row_data[16], row_data[17], row_data[18],
                                 row_data[19], row_data[20], row_data[21], row_data[22],
                                 row_data[23], row_data[24], row_data[25]))

            conn.commit()

    conn.close()


def copy_downloaded_file(fle):
    print("Copying {} to upload path".format(fle))
    shutil.copy2(os.path.join(g.excel_import_path, fle),
                 os.path.join(g.excel_import_path, 'upload', fle))


def update_excel_ship_date(fle):
    print("Updating ship date for {}".format(fle))
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    file_path = os.path.join(g.upload_file_path, fle)
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.active

    sql = "SELECT * FROM `records` WHERE `export_date` IS NULL AND `filename` = ? AND `kit_code` IS NOT NULL;"
    cursor.execute(sql, (fle,))
    results = cursor.fetchall()

    for rec in results:
        dt = datetime.datetime.now() + datetime.timedelta(days=2)
        ship_date = datetime.datetime.strftime(dt, '%m/%d/%Y')
        cell_data = ship_date
        ws.cell(row=rec['recno'] + 1, column=26, value=cell_data)

    wb.save(filename=os.path.join(g.upload_file_path, fle))
    conn.close()


def append_ship_date_to_clipboard(fle):
    print("Writing ship date to clipboard".format(fle))
    conn = sqlite3.connect(database=g.database)
    cursor = conn.cursor()

    sql = ("SELECT count(*) FROM `records` "
           "WHERE `export_date` IS NULL AND `filename` = ?"
           "GROUP BY `filename`;")
    cursor.execute(sql, (fle,))
    cnt = cursor.fetchone()
    conn.close()

    dt = datetime.datetime.now() + datetime.timedelta(days=2)
    ship_date = datetime.datetime.strftime(dt, '%m/%d/%Y')

    clip = (ship_date + "\n") * (cnt[0] - 1)
    clip = clip + ship_date
    pyperclip.copy(clip)


def write_outside_area_file_all():
    print("Writing outside area file")
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    select_fields = "`" + "`,`".join(g.outside_area_header) + "`"

    sql = f"SELECT {select_fields} FROM `records` WHERE `kit_code` IS NULL;"

    cursor.execute(sql)
    results = cursor.fetchall()

    with open(os.path.join(g.upload_file_path, "all_outside_area.csv"), 'w+', newline="") as s:
        csvw = csv.DictWriter(s, g.outside_area_header, delimiter=",", quoting=csv.QUOTE_ALL)
        csvw.writeheader()
        for rec in results:
            csvw.writerow(rec)

    conn.close()


def write_outside_area_file(fle):
    print("Writing outside area file for {}".format(fle))
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    select_fields = "`" + "`,`".join(g.outside_area_header) + "`"

    sql = (f"SELECT {select_fields} FROM `records` WHERE `export_date` IS NULL "
           "AND `filename` = ? AND `kit_code` IS NULL;")

    cursor.execute(sql, (fle,))
    results = cursor.fetchall()

    with open(os.path.join(g.upload_file_path, "{}_outside_area.csv".format(fle[:-5])), 'w+', newline="") as s:
        csvw = csv.DictWriter(s, g.outside_area_header, delimiter=",", quoting=csv.QUOTE_ALL)
        csvw.writeheader()
        for rec in results:
            csvw.writerow(rec)

    conn.close()


def write_letter_merge(fle):
    print("Writing letter merge for {}".format(fle))
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    sql = ("SELECT * FROM `records` WHERE `export_date` IS NULL "
           "AND `filename` = ? AND `kit_code` IS NOT NULL ORDER BY `kit_code`;")

    cursor.execute(sql, (fle,))
    results = cursor.fetchall()

    with open(os.path.join(g.letter_merge_path, "{}_merge.txt".format(fle[:-5])), 'w+', newline="") as s:
        csvw = csv.DictWriter(s, g.merge_letter_header, delimiter="\t")
        csvw.writeheader()
        for rec in results:
            name = str(rec['name']).strip()
            name = name.replace('Collateral Order', '').strip()
            w = {'Campaign': '',
                 'Individual_First_Name_1': name,
                 'Individual_Last_Name_1': '',
                 'Individual_First_Name_2': '',
                 'Individual_Last_Name_2': '',
                 'Address_1': rec['street_1'],
                 'Address_2': rec['street_2'],
                 'City': rec['city'],
                 'State': rec['state'],
                 'Zip': rec['zipcode'],
                 'County': rec['county'],
                 'Unique_ID': '',
                 'mid': '',
                 'art_code': '',
                 'kit': rec['kit_code']}

            csvw.writerow(w)

    conn.close()


def update_dates(fle):
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    dt = datetime.datetime.now() + datetime.timedelta(days=2)
    ship_date = datetime.datetime.strftime(dt, '%Y-%m-%d')

    print("Updating ship and export dates")

    sql1 = ("UPDATE `records` SET `export_date` = DATETIME('now', 'localtime') "
            "WHERE `filename` = ?;")

    sql2 = ("UPDATE `records` SET `ship_date` = ? "
            "WHERE `filename` = ?;")

    cursor.execute(sql1, (fle,))
    cursor.execute(sql2, (ship_date, fle,))

    conn.commit()
    conn.close()


def write_count_report(fle):
    print("Writing count report for {}".format(fle))
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    dt = datetime.datetime.now() + datetime.timedelta(days=2)
    ship_date = datetime.datetime.strftime(dt, '%Y-%m-%d')

    sql = ("SELECT `kit_code`, count(*) 'count' FROM `records` "
           "WHERE `export_date` IS NULL "
           "AND `filename` = ? AND `kit_code` IS NOT NULL "
           "GROUP BY `kit_code`;")

    cursor.execute(sql, (fle,))
    results = cursor.fetchall()

    cnt = 0
    with open(os.path.join(g.letter_merge_path, "Letter_Counts_{}.txt".format(ship_date)), 'w+', newline="") as s:
        csvw = csv.DictWriter(s, ['Ship Date', 'Kit Code', 'Count'], delimiter="\t")
        csvw.writeheader()
        for rec in results:
            w = {'Ship Date': ship_date, 'Kit Code': rec['kit_code'], 'Count': rec['count']}
            cnt += rec['count']
            csvw.writerow(w)

        csvw.writerow({'Ship Date': 'Total', 'Kit Code': '', 'Count': cnt})

    conn.commit()
    conn.close()


def move_file_to_complete(fle):
    print("Copying {} to complete path".format(fle))
    shutil.move(os.path.join(g.excel_import_path, fle),
                os.path.join(g.excel_import_path, 'complete', fle))


def update_kit_code(fle):
    print("Updating kit code for {}".format(fle))
    conn = sqlite3.connect(database=g.database)
    conn.row_factory = dict_factory
    cursor = conn.cursor()

    sql = "SELECT * FROM `records` WHERE `export_date` IS NULL AND `filename` = ?;"
    cursor.execute(sql, (fle,))
    results = cursor.fetchall()

    for result in results:
        rec_state = result['state']
        rec_county = str(result['county']).upper()
        kit_code = None

        if rec_state == 'IA' or rec_state == 'NE':
            kit_code = g.ia_ne_counties.get(f"{rec_state}_{rec_county}", None)
        elif rec_state == 'MN':
            kit_code = g.mn_counties.get(rec_county, None)

        sql = ("UPDATE `records` SET `kit_code` = ? "
               "WHERE `filename`||`recno` = ?||?;")

        cursor.execute(sql, (kit_code, result['filename'], result['recno'],))

    conn.commit()
    conn.close()


def final_message(fle):
    print(f"Processing complete for {fle}\nShip date copied to clipboard\n"
          f"Update {g.upload_file_path}\\{fle}, and upload to Dynamics\n"
          f"Create pdf letter files for production")


def init_db():

    conn = sqlite3.connect(database=g.database)
    cursor = conn.cursor()

    sql1 = ("CREATE table `records` ("
            "`filename` VARCHAR(100) NULL DEFAULT NULL,"
            "`recno` INT(10) NULL DEFAULT NULL,"
            "`import_date` DATETIME NULL DEFAULT NULL,"
            "`export_date` DATETIME NULL DEFAULT NULL,"
            "`kit_code` VARCHAR(20) NULL DEFAULT NULL,"
            "`collateral_kit_order` VARCHAR(100) NULL DEFAULT NULL," 
            "`row_checksum` VARCHAR(100) NULL DEFAULT NULL," 
            "`modified_on_do_not_mod` VARCHAR(100) NULL DEFAULT NULL," 
            "`name` VARCHAR(100) NULL DEFAULT NULL," 
            "`created_on` VARCHAR(100) NULL DEFAULT NULL," 
            "`created_by` VARCHAR(100) NULL DEFAULT NULL," 
            "`owner` VARCHAR(100) NULL DEFAULT NULL," 
            "`prospect_or_broker` VARCHAR(100) NULL DEFAULT NULL," 
            "`street_1` VARCHAR(100) NULL DEFAULT NULL," 
            "`street_2` VARCHAR(100) NULL DEFAULT NULL," 
            "`city` VARCHAR(100) NULL DEFAULT NULL," 
            "`full_state` VARCHAR(50) NULL DEFAULT NULL," 
            "`zipcode` VARCHAR(15) NULL DEFAULT NULL," 
            "`county` VARCHAR(100) NULL DEFAULT NULL," 
            "`state` VARCHAR(2) NULL DEFAULT NULL," 
            "`collateral_kit` VARCHAR(100) NULL DEFAULT NULL," 
            "`second_kit` VARCHAR(100) NULL DEFAULT NULL," 
            "`application_url` VARCHAR(100) NULL DEFAULT NULL," 
            "`comments` VARCHAR(100) NULL DEFAULT NULL," 
            "`vendor_unique_id_lead` VARCHAR(100) NULL DEFAULT NULL," 
            "`vendor_unique_id_contact` VARCHAR(100) NULL DEFAULT NULL," 
            "`quantity` VARCHAR(100) NULL DEFAULT NULL," 
            "`routing` VARCHAR(100) NULL DEFAULT NULL," 
            "`modified_by` VARCHAR(100) NULL DEFAULT NULL," 
            "`modified_on` VARCHAR(100) NULL DEFAULT NULL,"
            "`ship_date` DATE NULL DEFAULT NULL);")

    cursor.execute("DROP TABLE IF EXISTS `records`;")
    cursor.execute("VACUUM;")
    cursor.execute(sql1)

    conn.commit()
    conn.close()


def main():
    global g
    g = Global()
    g.initialize_config()
    # init_db()

    leads_files = [f for f in os.listdir(g.excel_import_path) if f[-4:].upper() == 'XLSX']
    for leads in leads_files:
        import_leads(leads)
        update_kit_code(leads)
        copy_downloaded_file(leads)
        append_ship_date_to_clipboard(leads)
        # update_excel_ship_date(leads)
        write_count_report(leads)
        write_letter_merge(leads)
        write_outside_area_file(leads)
        update_dates(leads)
        move_file_to_complete(leads)
        final_message(leads)


if __name__ == '__main__':
    main()
